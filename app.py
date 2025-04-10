import ee
import streamlit as st
import geopandas as gpd
import pandas as pd
import numpy as np
import datetime
import folium
from streamlit_folium import st_folium
from streamlit_oauth import OAuth2Component
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseUpload
import time
import requests
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
import gdown
import webbrowser
import matplotlib.pyplot as plt
import traceback
from PIL import Image
import io

# Configuração de layout
st.set_page_config(
    page_title="ZAP - Automatização",
    page_icon="🗺️",
    layout="wide",
    initial_sidebar_state="collapsed",
    menu_items={
        'Get help': 'https://www.mg.gov.br/agricultura/pagina/zoneamento-ambiental-e-produtive',
        'Report a Bug': "mailto:zap@agricultura.mg.gov.br"
    }
)

# Configurações para manter a sessão
if 'processing' not in st.session_state:
    st.session_state.processing = False
if 'processing_stage' not in st.session_state:  # 'agro', 'remoto', 'done'
    st.session_state.processing_stage = None
if 'tasks_remoto' not in st.session_state:
    st.session_state.tasks_remoto = []
if 'tasks_agro' not in st.session_state:
    st.session_state.tasks_agro = []
if 'completed_tasks' not in st.session_state:
    st.session_state.completed_tasks = 0
if 'resultados' not in st.session_state:
    st.session_state.resultados = None
if 'process_agro' not in st.session_state:
    st.session_state.process_agro = False
if 'select_all' not in st.session_state:
    st.session_state.select_all = False
if 'select_ibge' not in st.session_state:
    st.session_state.select_ibge = False
if 'agro_completed' not in st.session_state:
    st.session_state.agro_completed = False
if 'uploaded_file_name' not in st.session_state:
    st.session_state.uploaded_file_name = None

#Logo Sidebar e Sidebar
sidebar_logo = "https://i.postimg.cc/c4VZ0fQw/zap-logo.png"
main_body_logo = "https://i.postimg.cc/65qGpMc8/zap-logo-sb.png"
st.logo(sidebar_logo, size="large", icon_image=main_body_logo)

# Sidebar com links como markdown
with st.sidebar:
    st.markdown("## Navegação")
    
    # Links como markdown formatados como botões
    st.markdown("""
    <style>
        .sidebar-link {
            display: block;
            padding: 0.5rem 1rem;
            margin: 0.25rem 0;
            background-color: #f0f2f6;
            border-radius: 0.5rem;
            color: #333 !important;
            text-decoration: none !important;
            transition: all 0.3s;
            border: 1px solid #ddd;
        }
        .sidebar-link:hover {
            background-color: #e6e6e6;
            transform: translateX(3px);
        }
        
        /* Estilo dos botões para parecerem com os links */
        div[data-testid="stVerticalBlock"] > div[data-testid="stVerticalBlock"] > div > button {
            width: 100%;
            text-align: left;
            padding: 0.5rem 1rem;
            margin: 0.25rem 0;
            background-color: #f0f2f6 !important;
            border-radius: 0.5rem !important;
            color: #333 !important;
            border: 1px solid #ddd !important;
            box-shadow: none !important;
            transition: all 0.3s;
        }
        div[data-testid="stVerticalBlock"] > div[data-testid="stVerticalBlock"] > div > button:hover {
            background-color: #e6e6e6 !important;
            transform: translateX(3px);
        }
        div[data-testid="stVerticalBlock"] > div[data-testid="stVerticalBlock"] > div > button div p {
            font-weight: normal !important;
        }
    </style>
    """, unsafe_allow_html=True)

    # Link 1 - Sobre o ZAP
    st.markdown('<a href="https://www.mg.gov.br/agricultura/pagina/zoneamento-ambiental-e-produtive" class="sidebar-link" target="_blank">📘 Sobre o ZAP</a>', unsafe_allow_html=True)
    
    # Link 2 - Reportar Bug
    st.markdown('<a href="mailto:zap@agricultura.mg.gov.br" class="sidebar-link">🐞 Reportar um Bug</a>', unsafe_allow_html=True)
    
    # Botão para Política de Privacidade (abre modal)
    @st.dialog("Política de Privacidade", width="large")
    def show_privacy_policy():
        st.markdown("""
        *Última atualização: 31/03/2025*

        **1. Informações Gerais**
        - O aplicativo ZAP Automatização ("nós", "nosso" ou "aplicativo") é desenvolvido pela a Secretaria de Agricultura, Pecuária e Abastecimento de Minas Gerais como ferramenta de apoio ao Zoneamento Ambiental e Produtivo (ZAP). Esta política descreve como coletamos, usamos e protegemos suas informações.

        **2. Dados Coletados**
        - Autenticação Google: Utilizamos OAuth 2.0 para acessar serviços do Google (Earth Engine, Drive e Cloud) com seu consentimento explícito.
        - Arquivos GeoJSON: Arquivos geográficos enviados para processamento são armazenados temporariamente apenas durante a sessão.
        - Dados de Uso: Registramos operações realizadas para fins de auditoria e melhoria do serviço.

        **3. Uso dos Dados**
        Os dados coletados são usados exclusivamente para:
        - Processamento de informações geográficas
        - Geração de relatórios e produtos do ZAP
        - Melhoria contínua do aplicativo

        **4. Compartilhamento de Dados**
        Não compartilhamos seus dados pessoais com terceiros, exceto:
        - Quando exigido por lei
        - Para prestação de serviços Google necessários ao funcionamento do aplicativo

        **5. Segurança**
        - Implementamos medidas técnicas e organizacionais para proteger seus dados, incluindo:
        - Autenticação em dois fatores recomendada
        - Acesso restrito a pessoal autorizado
        - Criptografia de dados em trânsito

        **6. Seus Direitos**
        Você pode:
        - Revogar o acesso à sua conta Google a qualquer momento
        - Solicitar acesso aos dados armazenados
        - Requerer a exclusão de seus dados

        **7. Alterações na Política**
        - Esta política pode ser atualizada periodicamente. Alterações significativas serão comunicadas aos usuários.
        """)
    
    # Botão para Aspectos Legais (abre modal)
    @st.dialog("Termos de Serviço", width="large")
    def show_legal_terms():
        st.markdown("""
        *Última atualização: 31/03/2025*

        **1. Aceitação dos Termos**
        - Ao utilizar o aplicativo ZAP Automatização, você concorda com estes Termos de Serviço.

        **2. Uso Autorizado**
        O aplicativo destina-se exclusivamente a:
        - Técnicos e gestores públicos envolvidos com o ZAP
        - Usuários autorizados pela Secretaria de Agricultura de MG

        **3. Responsabilidades do Usuário**
        Você concorda em:
        - Fornecer apenas informações precisas e atualizadas
        - Não utilizar o aplicativo para fins ilegais
        - Manter suas credenciais de acesso em sigilo

        **4. Limitações**
        O aplicativo não garante:
        - Disponibilidade contínua ou ininterrupta
        - Precisão absoluta dos resultados processados
        - Compatibilidade com todos os sistemas ou dispositivos

        **5. Propriedade Intelectual**
        - Todo o conteúdo e funcionalidades do aplicativo são propriedade do Governo de Minas Gerais e estão protegidos por leis de propriedade intelectual.

        **6. Isenção de Responsabilidade**
        Não nos responsabilizamos por:
        - Danos resultantes do uso inadequado do aplicativo
        - Perda de dados devido a falhas técnicas
        - Conteúdo gerado por terceiros

        **7. Rescisão**
        - Reservamos o direito de encerrar o acesso ao aplicativo a qualquer usuário que violar estes Termos.

        **8. Legislação Aplicável**
        - Estes Termos são regidos pelas leis brasileiras e quaisquer disputas serão resolvidas no foro da Comarca de Belo Horizonte/MG.

        ## Contato
        Para questões sobre privacidade ou termos de serviço:
        - Email: zap@agricultura.mg.gov.br
        - Site: [ZAP Minas Gerais - SEAPA](https://www.mg.gov.br/agricultura/pagina/zoneamento-ambiental-e-produtive)

        ## Outros Links:
        - [Aspectos Legais e Responsabilidades (Governo de MG)](https://www.mg.gov.br/pagina/aspectos-legais-e-responsabilidades)
        - [Política de Privacidade (SEAPA-MG)](https://www.mg.gov.br/agricultura/pagina/politica-de-privacidade)
        """)
    
    # Botões que acionam os diálogos
    if st.button("🔒 Política de Privacidade", key="privacy_button"):
        show_privacy_policy()
    
    if st.button("⚖️ Termos de Serviço", key="legal_button"):
        show_legal_terms()

    st.markdown("---")
    st.markdown("### Versão 1.0")
    st.caption("Desenvolvido para a 5ª edição do ZAP")
    st.caption("Secretaria de Agricultura, Pecuária e Abastecimento de Minas Gerais")
    
# Logo e título centralizado
col1, col2, col3 = st.columns([1,2,1])
with col2:
    st.image("https://i.postimg.cc/c4VZ0fQw/zap-logo.png", width=400)

# Título do aplicativo
st.title("Automatização de Obtenção de Dados para o Zoneamento Ambiental e Produtivo")

# CSS customizado para os cards
st.markdown("""
<style>
    .custom-card {
        background-color: #242434;
        border-radius: 15px;
        padding: 1.5rem;
        margin: 1rem 0;
        box-shadow: 0 4px 6px rgba(0, 0, 0, 0.05);
        border-left: 4px solid #2e7d32;
    }
    .custom-card h3 {
        color: #2e7d32;
        margin-top: 0;
    }
    .custom-card ul {
        padding-left: 1.2rem;
    }
    .custom-card a {
        color: #1a5a96 !important;
        font-weight: 500;
    }
        .stFileUploader > label > div:first-child {
        font-weight: bold;
        color: #ff4b4b;
    }
    .stFileUploader > label > div:nth-child(2) {
        font-size: 0.8em;
        color: #777;
    }
    /* Estilo específico para os botões dentro do .custom-card */
    .custom-card .stButton button {
        width: 100%;
        text-align: center;
        padding: 0.5rem 1rem;
        margin: 0.75rem 0;
        background-color: #2e7d32 !important;
        color: white !important;
        border-radius: 8px !important;
        border: none !important;
        transition: all 0.3s;
    }
    .custom-card .stButton button:hover {
        background-color: #1e5e22 !important;
        transform: translateY(-2px);
        box-shadow: 0 2px 4px rgba(0,0,0,0.2);
    }
    .card-buttons-container {
        display: flex;
        gap: 1rem;
        margin-top: 1.5rem;
    }
    @media (max-width: 768px) {
        .card-buttons-container {
            flex-direction: column;
            gap: 0.5rem;
        }
    }
</style>
""", unsafe_allow_html=True)

# Card Política de Privacidade e Termo de Serviço
st.markdown("""
<div class="custom-card">
    <h3>🔒 Política de Privacidade e Termos de Serviço ⚖️</h3>
    <p>Antes de iniciar, leia nossa Política de Privacidade e nossos Termos de Serviço.</p>
    <div class="card-buttons-container">
""", unsafe_allow_html=True)

# Botões dentro do card (usando columns para layout responsivo)

if st.button("🔒 Política de Privacidade", 
            key="card_privacy_button",
            help="Clique para ver nossa política de privacidade"):
    show_privacy_policy()

if st.button("⚖️ Termos de Serviço", 
            key="card_legal_button",
            help="Clique para ver os termos de serviço"):
    show_legal_terms()

st.markdown("</div></div>", unsafe_allow_html=True)

# Card 1 - Sobre o ZAP
st.markdown("""
<div class="custom-card">
<h3>🌱 Sobre o ZAP</h3>
O Zoneamento Ambiental e Produtivo (ZAP) é um instrumento de planejamento e gestão territorial para o uso sustentável dos recursos naturais pela atividade agrossilvipastoril no estado de Minas Gerais, instituído pelo Decreto Estadual nº 46.650/2014.

<h3>🗺️ Produtos Básicos</h3>
<ul>
<li>Mapeamento da cobertura e terra</li>
<li>Índice de Demanda Hídrica Superficial (IDHS)</li>
<li>Potencial de Uso Conservacionista (PUC)</li>
</ul>

O ZAP busca disponibilizar informações detalhadas sobre o meio natural e produtivo por sub-bacia hidrográfica.

<h3>🔄 Evolução da Metodologia</h3>
Desenvolvida inicialmente pela Semad e Seapa em 2014, a metodologia do ZAP está atualmente na 5ª edição (2025). O Comitê Gestor do ZAP é a instância consultiva e deliberativa da ferramenta.

<h3>🤝 Integração com Outros Instrumentos</h3>
<ul>
<li>Indicadores de Sustentabilidade em Agroecossistemas (ISAs)</li>
<li>Planos de Adequação Socioeconômica e Ambiental (PASEAs)</li>
<li>Cadastro Ambiental Rural (CAR)</li>
<li>Entre outros</li>
</ul>

🔗 <a href="https://www.mg.gov.br/agricultura/pagina/zoneamento-ambiental-e-produtive" target="_blank">Mais informações no Site do Governo de MG</a>
</div>
""", unsafe_allow_html=True)

# Card 2 - Sobre a Ferramenta
st.markdown("""
<div class="custom-card">
<h3>🛠️ Sobre esta Ferramenta</h3>
Esta ferramenta automatiza a obtenção de produtos e bases para os produtos utilizados no ZAP para a 5ª edição da metodologia.

<h3>🔑 Requisitos</h3>
<ul>
<li>Conexão com conta Google (para Earth Engine, Cloud Service e Drive)</li>
<li>Projeto na Google Cloud Service com acesso à API da Earth Engine</li>
<li>Arquivo GeoJSON da bacia hidrográfica (preferencialmente em UTM)</li>
</ul>
</div>
""", unsafe_allow_html=True)

# Divisão visual
st.markdown("---")

# 1. Configuração inicial e autenticação
if 'google_oauth' in st.secrets:
    CLIENT_ID = st.secrets['google_oauth']['client_id']
    CLIENT_SECRET = st.secrets['google_oauth']['client_secret']
    REDIRECT_URI = st.secrets['google_oauth']['redirect_uris']
else:
    st.error("Configurações do OAuth2 não encontradas no secrets.toml.")
    st.stop()

AUTHORIZE_URL = "https://accounts.google.com/o/oauth2/auth"
TOKEN_URL = "https://oauth2.googleapis.com/token"
REFRESH_TOKEN_URL = "https://oauth2.googleapis.com/token"
REVOKE_TOKEN_URL = "https://oauth2.googleapis.com/revoke"

SCOPES = [
    "https://www.googleapis.com/auth/earthengine",
    "https://www.googleapis.com/auth/cloud-platform",
    "https://www.googleapis.com/auth/drive.file",
]
SCOPE = " ".join(SCOPES)

oauth2 = OAuth2Component(CLIENT_ID, CLIENT_SECRET, AUTHORIZE_URL, TOKEN_URL, REFRESH_TOKEN_URL, REVOKE_TOKEN_URL)

# 2. Dicionário de produtos (completo)
DICIONARIO_PRODUTOS = {
    #PAM (Produção Agrícola Municipal)
    'abacate': ('Abacate', '#568203'),
    'abacaxi': ('Abacaxi', '#FEE347'),
    'algodaa': ('Algodão arbóreo', '#D3D3D3'),
    'algodah': ('Algodão herbáceo', '#E6E6E6'),
    'alho': ('Alho', '#B19CD9'),
    'amendoi': ('Amendoim', '#D2B48C'),
    'arroz': ('Arroz', '#FFF8DC'),
    'aveia': ('Aveia', '#D0C487'),
    'azeiton': ('Azeitona', '#808000'),
    'acai': ('Açaí', '#4B0082'),
    'banana': ('Banana', '#FFE135'),
    'batatad': ('Batata-doce', '#D2691E'),
    'batatai': ('Batata-inglesa', '#EED9C4'),
    'borrach': ('Borracha', '#696969'),
    'cacau': ('Cacau', '#4B3621'),
    'cafeara': ('Café Arábica', '#6F4E37'),
    'cafecan': ('Café Canephora', '#8B4513'),
    'cafetot': ('Café Total', '#A0522D'),
    'cana': ('Cana-de-açúcar', '#3A913F'),
    'caqui': ('Caqui', '#FFA07A'),
    'castcaj': ('Castanha de caju', '#FFD700'),
    'cebola': ('Cebola', '#9F4576'),
    'cenoura': ('Cenoura', '#ED9121'),
    'centeio': ('Centeio', '#D2B48C'),
    'cevada': ('Cevada', '#F0E68C'),
    'chaind': ('Chá-da-índia', '#D8BFD8'),
    'cocobai': ('Coco-da-baía', '#F5DEB3'),
    'dende': ('Dendê', '#6B8E23'),
    'ervamat': ('Erva-mate', '#556B2F'),
    'ervilha': ('Ervilha', '#90EE90'),
    'fava': ('Fava', '#8FBC8F'),
    'feijao': ('Feijão', '#8B4513'),
    'figo': ('Figo', '#9400D3'),
    'fumo': ('Fumo', '#708090'),
    'girass': ('Girassol', '#FFD700'),
    'goiaba': ('Goiaba', '#FF6347'),
    'guarana': ('Guaraná', '#8B0000'),
    'juta': ('Juta', '#F5DEB3'),
    'laranja': ('Laranja', '#FFA500'),
    'limao': ('Limão', '#F5F5DC'),
    'linho': ('Linho', '#FAF0E6'),
    'mamona': ('Mamona', '#8B008B'),
    'mamao': ('Mamão', '#FF6347'),
    'mandioc': ('Mandioca', '#F5DEB3'),
    'manga': ('Manga', '#FF8243'),
    'maracuj': ('Maracujá', '#9370DB'),
    'marmelo': ('Marmelo', '#DAA520'),
    'maca': ('Maçã', '#FF0800'),
    'melanci': ('Melancia', '#FC6C85'),
    'melao': ('Melão', '#FDBE02'),
    'milho': ('Milho', '#F2D024'),
    'morango': ('Morango', '#D53032'),
    'noz': ('Noz', '#800000'),
    'palmito': ('Palmito', '#6B8E23'),
    'pera': ('Pera', '#D1E231'),
    'pimrein': ('Pimenta-do-reino', '#A0522D'),
    'pessego': ('Pêssego', '#FFDAB9'),
    'rami': ('Rami', '#6B8E23'),
    'sisal': ('Sisal', '#F5F5DC'),
    'soja': ('Soja', '#D4A017'),
    'sorgo': ('Sorgo', '#D2B48C'),
    'tangeri': ('Tangerina', '#F28500'),
    'tomate': ('Tomate', '#FF6347'),
    'trigo': ('Trigo', '#F5DEB3'),
    'tritica': ('Triticale', '#D8BFD8'),
    'tungue': ('Tungue', '#8B4513'),
    'urucum': ('Urucum', '#B22222'),
    'uva': ('Uva', '#6F2DA8'),
    #PPM (Pecuária Municipal)
    'bovino': ('Bovino', '#8B4513'),
    'bubalin': ('Bubalino', '#A0522D'),
    'caprino': ('Caprino', '#CD853F'),
    'codorna': ('Codornas', '#F5DEB3'),
    'equino': ('Equino', '#D2691E'),
    'galin': ('Galináceos', '#FFD700'),
    'ovino': ('Ovino', '#F0E68C'),
    'suino': ('Suíno', '#FFC0CB'),
    'bichsed': ('Casulos do bicho-da-seda', '#F5F5DC'),
    'leite': ('Leite', '#ADD8E6'),
    'la': ('Lã', '#E6E6FA'),
    'mel': ('Mel', '#DAA520'),
    'ovocod': ('Ovos de codorna', '#F5DEB3'),
    'ovogal': ('Ovos de galinha', '#FFD700'),
    'alevino': ('Alevinos', '#87CEEB'),
    'camarao': ('Camarão', '#E2725B'),
    'carpa': ('Carpa', '#FFA500'),
    'curimat': ('Curimatã', '#4682B4'),
    'dourado': ('Dourado', '#FFD700'),
    'jatuara': ('Jatuarana', '#1E90FF'),
    'lambari': ('Lambari', '#00BFFF'),
    'camlarv': ('Larvas de camarão', '#FF6347'),
    'matrinx': ('Matrinxã', '#4169E1'),
    'mexilh': ('Mexilhões', '#778899'),
    'outpeix': ('Outros peixes', '#87CEEB'),
    'pacu': ('Pacu', '#00BFFF'),
    'piau': ('Piau', '#1E90FF'),
    'pintado': ('Pintado', '#483D8B'),
    'pirapi': ('Pirapitinga', '#4682B4'),
    'piraruc': ('Pirarucu', '#1E90FF'),
    'semmol': ('Sementes de moluscos', '#F5DEB3'),
    'tambacu': ('Tambacu', '#6495ED'),
    'tambaqu': ('Tambaqui', '#4169E1'),
    'tilapia': ('Tilápia', '#4682B4'),
    'traira': ('Traíra', '#2F4F4F'),
    'truta': ('Truta', '#ADD8E6'),
    'tucuna': ('Tucunaré', '#000080'),
    #PEVS (Produção da Extração Vegetal e Silvicultura)
    'eucalip': ('Eucalipto', '#228B22'),
    'outesp': ('Outras espécies', '#D3D3D3'),
    'pinus': ('Pinus', '#2E8B57'),
    'carveg': ('Carvão vegetal', '#36454F'),
    'lenha': ('Lenha', '#A0522D'),
    'madtor': ('Madeira em tora', '#8B0000'),
    'outprod': ('Outros produtos', '#A9A9A9')
}

# 3. URLs das tabelas (convertidas para links diretos do Google Drive)
TABELAS_AGRO = {
    'PAM_Quantidade_produzida_14-23': 'https://drive.google.com/uc?id=19BaNA96nXA4gtkmF_nwSQFdxA5UEBmmx',
    'PAM_Valor_da_producao_14-23': 'https://drive.google.com/uc?id=1A9o-eEiXpPMWOyCtE4m2jwYaovRy9bv9',
    'PPM_Efetivo_dos_rebanhos_14-23': 'https://drive.google.com/uc?id=1VTNqLYXi5AjiWCZDu2cUfbmVzwYjbLrY',
    'PPM_Prod_origem_animal_14-23': 'https://drive.google.com/uc?id=18I1Yr7qsICf8hBtBawkmG9Wes5Hd2hBz',
    'PPM_Valor_da_producao_prod_animal_14-23': 'https://drive.google.com/uc?id=1s-9uSiVOxZJLgIKVP8ZI8rCo99DgiEIf',
    'PPM_Producao_aquicultura_14-23': 'https://drive.google.com/uc?id=16VeRUfYvGgj2_swg_g671uJ_5I2QPpo2',
    'PPM_Valor_producao_aquicultura_14-23': 'https://drive.google.com/uc?id=19-yrafwVj0ZOPiqbwhqX1Ho3Gwr1GoIA',
    'PEVS_Area_silv_14-23': 'https://drive.google.com/uc?id=10uwm4SgvYKDzTpi2jlPirjzPcL_5PTCB',
    'PEVS_Qnt_prod_silv_14-23': 'https://drive.google.com/uc?id=1qIHRhddxGV8WtEEt0lJcaxnUpKjF1MBK',
    'PEVS_Valor_prod_silv_14-23': 'https://drive.google.com/uc?id=1BzPQy5pFNrqgC_9gHCUDO7Sy4O-t6nrA',
    'IBGE_Municipios_ZAP': 'https://drive.google.com/uc?id=1skVkA0cN3TVlJThvqsilWwO2SGLY-joi'
}

# 4. Funções auxiliares
def load_geojson(file):
    try:
        # Verificação de tamanho (1 MB)
        MAX_FILE_SIZE_KB = 1024
        if file.size > MAX_FILE_SIZE_KB * 1024:
            st.error(f"Tamanho do arquivo ({file.size/1024:.1f} KB) excede o limite de {MAX_FILE_SIZE_KB} KB")
            return None, None

        gdf = gpd.read_file(file)
        
        # Verifica número de features
        if len(gdf) > 1:
            st.error("Erro: O arquivo deve conter APENAS UMA feature (polígono/multipolígono).")
            return None, None
            
        # Validações de geometria
        if gdf.geometry.is_empty.any():
            st.error("Erro: O arquivo contém geometrias vazias.")
            return None, None
            
        if not all(gdf.geometry.geom_type.isin(['Polygon', 'MultiPolygon'])):
            st.error("Erro: Apenas polígonos ou multipolígonos são aceitos.")
            return None, None

        # Verificação do CRS (APENAS SIRGAS 2000)
        CRS_OBRIGATORIO = 'EPSG:4674'
        if gdf.crs is None:
            st.error(f"Erro: O arquivo não possui CRS definido. O CRS deve ser {CRS_OBRIGATORIO} (SIRGAS 2000).")
            return None, None
            
        if str(gdf.crs).upper() != CRS_OBRIGATORIO:
            st.error(f"Erro: CRS {gdf.crs} não permitido. O arquivo deve estar em {CRS_OBRIGATORIO} (SIRGAS 2000).")
            return None, None

        # Correção de geometrias (buffer 0 se necessário)
        if not gdf.geometry.is_valid.all():
            gdf['geometry'] = gdf.geometry.buffer(0)
        
        # Visualização DIRETAMENTE no SIRGAS 2000 (Folium aceita coordenadas equivalentes)
        centroid = gdf.geometry.centroid
        m = folium.Map(
            location=[centroid.y.mean(), centroid.x.mean()],  # Coordenadas serão interpretadas como WGS84
            zoom_start=10,
            tiles='CartoDB positron'
        )
        
        folium.GeoJson(
            gdf.geometry.iloc[0],
            style_function=lambda x: {'fillColor': '#4daf4a', 'color': '#377eb8'}
        ).add_to(m)
            
        st_folium(m, width=600, height=400)
        st.success(f"CRS do arquivo validado: {gdf.crs} (SIRGAS 2000)")
        
        # Retorna geometria no SIRGAS 2000
        return ee.Geometry(gdf.geometry.iloc[0].__geo_interface__), CRS_OBRIGATORIO
        
    except Exception as e:
        st.error(f"Erro crítico ao carregar GeoJSON: {str(e)}")
        return None, None

def reprojetarImagem(imagem, epsg, escala):
    return imagem.reproject(crs=f"EPSG:{epsg}", scale=escala)

def exportarImagem(imagem, nome_prefixo, nome_sufixo, escala, regiao, nome_bacia_export, pasta="zap"):
    try:
        nome_arquivo = f"{nome_prefixo}{nome_bacia_export}{nome_sufixo}"
        task = ee.batch.Export.image.toDrive(
            image=imagem,
            description=nome_arquivo,
            folder=pasta,
            fileNamePrefix=nome_arquivo,
            scale=escala,
            region=regiao,
            fileFormat='GeoTIFF',
            maxPixels=1e13,
        )
        task.start()
        st.success(f"Exportação {nome_arquivo} iniciada. Verifique seu Google Drive na pasta '{pasta}'.")
        return task
    except Exception as e:
        st.error(f"Erro ao exportar {nome_arquivo} para o Google Drive: {e}")
        return None

def check_task_status(task):
    try:
        status = task.status()
        state = status["state"]
        if state == "COMPLETED":
            return state
        elif state == "RUNNING":
            return state
        elif state == "FAILED":
            return state
        else:
            return state
    except Exception as e:
        st.error(f"Erro ao verificar o status da tarefa: {e}")
        return None

# 5. Funções para processamento dos dados agro
def processar_municipios(geometry, nome_bacia_export):
    try:
        # Carregar municípios de MG (do Earth Engine)
        municipios_mg = ee.FeatureCollection("projects/ee-zapmg/assets/mg-municipios")
        
        # Calcular área da bacia
        area_bacia = geometry.area()
        
        # Função para calcular interseção
        def calcular_intersecao(feature):
            intersecao = feature.geometry().intersection(geometry, 1)
            area_intersecao = intersecao.area()
            percentual = area_intersecao.divide(area_bacia).multiply(100)
            
            return feature.set({
                'area_intersecao_ha': area_intersecao.divide(10000),
                'percentual_na_bacia': percentual,
                'area_municipio_ha': feature.geometry().area().divide(10000),
                'area_bacia_ha': area_bacia.divide(10000)
            })
        
        # Processar todos os municípios que intersectam
        municipios_processados = municipios_mg.filterBounds(geometry).map(calcular_intersecao)
        
        # Filtrar municípios com mais de 20% de representatividade
        municipios_selecionados = municipios_processados.filter(ee.Filter.gte('percentual_na_bacia', 20))
        
        # Converter para Pandas DataFrame
        features = municipios_selecionados.getInfo()['features']
        dados_municipios = []
        
        for feature in features:
            props = feature['properties']
            props['geocodigo'] = int(props['geocodigo'])  # Garantir que é inteiro
            dados_municipios.append(props)
        
        df_municipios = pd.DataFrame(dados_municipios)
        
        if not df_municipios.empty:
            df_municipios = df_municipios.sort_values('percentual_na_bacia', ascending=False)
            
            # Formatando as colunas numéricas
            df_municipios['area_intersecao_ha'] = df_municipios['area_intersecao_ha'].round(2)
            df_municipios['percentual_na_bacia'] = df_municipios['percentual_na_bacia'].round(2)
            df_municipios['area_municipio_ha'] = df_municipios['area_municipio_ha'].round(2)
            
            # Criando DataFrame para exibição
            df_display = df_municipios[['nome', 'area_intersecao_ha', 'percentual_na_bacia']].copy()
            df_display.columns = ['Município', 'Área na Bacia (ha)', 'Representatividade (%)']
            
            st.success(f"{len(df_municipios)} município(s) selecionado(s) com mais de 20% de área na bacia")
            
            # Mostrar tabela detalhada
            st.write("### Detalhes dos Municípios")
            st.dataframe(df_display.sort_values('Representatividade (%)', ascending=False))
            
            # Adicionar métricas resumidas
            col1, col2, col3 = st.columns(3)
            with col1:
                st.metric("Municípios selecionados", len(df_municipios))
            with col2:
                st.metric("Área total na bacia (ha)", 
                         round(df_municipios['area_intersecao_ha'].sum(), 2))
            with col3:
                st.metric("Representatividade média (%)", 
                         round(df_municipios['percentual_na_bacia'].mean(), 2))
            
            return df_municipios
        else:
            st.warning("Nenhum município com mais de 20% de área na bacia foi encontrado.")
            return None
            
    except Exception as e:
        st.error(f"Erro ao processar municípios: {e}")
        return None

def baixar_tabela(url):
    try:
        output = BytesIO()
        gdown.download(url, output, quiet=True)
        output.seek(0)
        return pd.read_csv(output)
    except Exception as e:
        st.error(f"Erro ao baixar tabela: {e}")
        return None

def processar_tabelas_agro(geocodigos):
    resultados = {}
    
    # Processar todas as tabelas, incluindo IBGE
    for nome_tabela, url in TABELAS_AGRO.items():
        df = baixar_tabela(url)
        if df is None:
            continue
            
        # Converter geocodigo para inteiro
        df['geocodigo'] = df['geocodigo'].astype(int)
        
        # Filtrar municípios selecionados
        df_filtrado = df[df['geocodigo'].isin(geocodigos)]
        
        if df_filtrado.empty:
            continue
            
        # Tratamento especial para tabela IBGE
        if nome_tabela == 'IBGE_Municipios_ZAP':
            # Remover colunas indesejadas (mas manter 'Municípios')
            colunas_remover = [col for col in ['.geo', 'system:index'] if col in df_filtrado.columns]
            if colunas_remover:
                df_filtrado = df_filtrado.drop(columns=colunas_remover)
            
            # Renomear colunas conforme solicitado
            renomear = {
                'População ocupada': 'População ocupada {%}',
                'Densidade demográfica': 'Densidade demográfica (hab/km²)',
                'Esgotamento sanitário adequado': 'Esgotamento sanitário adequado {%}',
                'Mortalidade Infantil': 'Mortalidade Infantil {%}',
                'Taxa de escolarização de 6 a 14 anos de idade': 'Taxa de escolarização de 6 a 14 anos de idade {%}',
                'Urbanização de vias públicas': 'Urbanização de vias públicas {%}',
                'Arborização de vias públicas': 'Arborização de vias públicas {%}'
            }
            df_filtrado = df_filtrado.rename(columns=renomear)
            
            # Definir a ordem das colunas (incluindo 'Municípios')
            ordem_colunas = [
                'Municípios',
                'geocodigo',
                'Gentílico',
                'Bioma predominante',
                'Área (km²)',
                'População no último censo',
                'População ocupada {%}',
                'Densidade demográfica (hab/km²)',
                'PIB per capita',
                'Salário médio mensal dos trabalhadores formais',
                'Receitas',
                'Despesas',
                'Esgotamento sanitário adequado {%}',
                'Estabelecimentos de Saúde SUS',
                'Mortalidade Infantil {%}',
                'Taxa de escolarização de 6 a 14 anos de idade {%}',
                'Urbanização de vias públicas {%}',
                'Arborização de vias públicas {%}',
                'Índice de Desenvolvimento Humano Municipal (IDHM)'
            ]
            
            # Manter apenas as colunas que existem no DataFrame
            ordem_colunas = [col for col in ordem_colunas if col in df_filtrado.columns]
            
            # Reordenar as colunas e transpor
            df_final = df_filtrado[ordem_colunas].set_index('Municípios').T
            df_final.index.name = 'Indicador'
            resultados[nome_tabela] = df_final
            continue
            
        # Para outras tabelas, criar uma planilha com top 10 produtos por município
        municipios_dfs = {}
        for _, row in df_filtrado.iterrows():
            municipio = row['nome']
            geocodigo = row['geocodigo']
            
            # Identificar colunas de anos (terminadas com 2 dígitos)
            colunas_ano = [col for col in row.index if col[-2:].isdigit() and col not in ['geocodigo', 'nome']]
            
            # Agrupar por produto (prefixo antes do ano)
            produtos = {}
            for col in colunas_ano:
                produto = col[:-2]
                ano = col[-2:]
                valor = row[col]
                
                if produto not in produtos:
                    produtos[produto] = {}
                produtos[produto][ano] = valor
            
            # Converter para DataFrame e pegar top 10 produtos com maior valor em 2023
            df_produtos = pd.DataFrame.from_dict(produtos, orient='index')
            
            # Ordenar por 2023 (se existir) ou pelo último ano disponível
            if '23' in df_produtos.columns:
                df_produtos = df_produtos.sort_values('23', ascending=False)
            else:
                ultimo_ano = sorted(df_produtos.columns)[-1]
                df_produtos = df_produtos.sort_values(ultimo_ano, ascending=False)
            
            # Pegar top 10 e traduzir nomes
            top_10 = df_produtos.head(10)
            top_10.index = [DICIONARIO_PRODUTOS.get(p, p) for p in top_10.index]
            
            # Adicionar município como coluna
            top_10 = top_10.reset_index()
            top_10.columns = ['Produto'] + [f'20{ano}' for ano in top_10.columns[1:]]
            
            municipios_dfs[municipio] = top_10
        
        resultados[nome_tabela] = municipios_dfs
    
    return resultados

def criar_grafico_unico_municipio(df_municipio, municipio, tipo_dado, tabela_origem):
    """Cria gráficos com a evolução dos produtos para um município, com formatação específica."""
    try:
        if len(df_municipio) == 0:
            return None

        # Dicionário de títulos personalizados
        titulos_por_tabela = {
            'PAM_Quantidade_produzida_14-23': 'PAM - Evolução da Quantidade Produzida',
            'PAM_Valor_da_producao_14-23': 'PAM - Evolução do Valor da Produção',
            'PPM_Efetivo_dos_rebanhos_14-23': 'PPM - Evolução do Efetivo dos Rebanhos',
            'PPM_Prod_origem_animal_14-23': 'PPM - Evolução da Quantidade de Produtos de Origem Animal',
            'PPM_Valor_da_producao_prod_anim': 'PPM - Evolução do Valor/Receita dos Produtos de Origem Animal',
            'PPM_Producao_aquicultura_14-23': 'PPM - Evolução da Quantidade Produzida na Aquicultura',
            'PPM_Valor_producao_aquicultura_': 'PPM - Evolução do Valor/Receita da Produção na Aquicultura',
            'PEVS_Area_silv_14-23': 'PEVS - Evolução da Área de Silvicultura',
            'PEVS_Qnt_prod_silv_14-23': 'PEVS - Evolução da Quantidade Produzida na Silvicultura',
            'PEVS_Valor_prod_silv_14-23': 'PEVS - Evolução do Valor da Produção na Silvicultura'
        }

        # Dicionário de unidades e ajustes de escala
        unidades_config = {
            'PAM_Quantidade_produzida_14-23': {'unidade': 'Mil Toneladas', 'divisor': 1000},
            'PAM_Valor_da_producao_14-23': {'unidade': 'Mil Reais', 'divisor': 1},
            'PPM_Efetivo_dos_rebanhos_14-23': {'unidade': 'Cabeças', 'divisor': 1},
            'PPM_Prod_origem_animal_14-23': {
                'unidades_especificas': {
                    'leite': {'unidade': 'Mil Litros', 'divisor': 1},
                    'ovogal': {'unidade': 'Mil Dúzias', 'divisor': 1},
                    'ovocod': {'unidade': 'Mil Dúzias', 'divisor': 1},
                    'mel': {'unidade': 'Quilogramas', 'divisor': 1},
                    'bichsed': {'unidade': 'Quilogramas', 'divisor': 1}
                },
                'default': {'unidade': 'Unidade', 'divisor': 1}
            },
            'PPM_Valor_da_producao_prod_anim': {'unidade': 'Mil Reais', 'divisor': 1},
            'PPM_Producao_aquicultura_14-23': {'unidade': 'Quilogramas', 'divisor': 1},
            'PPM_Valor_producao_aquicultura_': {'unidade': 'Mil Reais', 'divisor': 1},
            'PEVS_Area_silv_14-23': {'unidade': 'Hectares', 'divisor': 1, 'ylabel': 'Área'},
            'PEVS_Qnt_prod_silv_14-23': {
                'unidades_especificas': {
                    'carveg': {'unidade': 'Toneladas', 'divisor': 1},
                    'outprod': {'unidade': 'Toneladas', 'divisor': 1},
                    'lenha': {'unidade': 'Metros Cúbicos', 'divisor': 1},
                    'madtor': {'unidade': 'Metros Cúbicos', 'divisor': 1}
                },
                'default': {'unidade': 'Unidade', 'divisor': 1}
            },
            'PEVS_Valor_prod_silv_14-23': {'unidade': 'Mil Reais', 'divisor': 1}
        }

        # Lista de tabelas com gráficos únicos (que precisam de altura maior)
        tabelas_com_grafico_unico = [
            'PAM_Quantidade_produzida_14-23',
            'PAM_Valor_da_producao_14-23',
            'PPM_Efetivo_dos_rebanhos_14-23',
            'PPM_Valor_da_producao_prod_anim',
            'PPM_Producao_aquicultura_14-23',
            'PPM_Valor_producao_aquicultura_',
            'PEVS_Area_silv_14-23',
            'PEVS_Valor_prod_silv_14-23'
        ]

        # Obter configurações da tabela
        config = unidades_config.get(tabela_origem, {'unidade': 'Unidade', 'divisor': 1})
        titulo_base = titulos_por_tabela.get(tabela_origem, f"Evolução {tipo_dado}")

        # Função para extrair o nome e chave do produto
        def get_produto_info(produto_info):
            if isinstance(produto_info, tuple):
                produto_nome = produto_info[0]
                produto_key = produto_nome.lower().replace(' ', '').replace('-', '').replace('_', '')
                cor = produto_info[1]
            else:
                produto_nome = produto_info
                produto_key = produto_nome.lower().replace(' ', '').replace('-', '').replace('_', '')
                cor_info = DICIONARIO_PRODUTOS.get(produto_key, ('', '#A9A9A9'))
                cor = cor_info[1] if isinstance(cor_info, tuple) else '#A9A9A9'
            return produto_nome, produto_key, cor

        # Função para determinar a unidade de cada produto
        def get_unidade_produto(produto_key):
            if 'unidades_especificas' in config:
                for k, v in config['unidades_especificas'].items():
                    if produto_key.startswith(k):
                        return v['unidade'], v['divisor']
            return config.get('unidade', 'Unidade'), config.get('divisor', 1)

        # Agrupar produtos por unidade de medida
        grupos = {}
        for _, row in df_municipio.iterrows():
            produto_info = row['Produto']
            produto_nome, produto_key, _ = get_produto_info(produto_info)
            unidade, divisor = get_unidade_produto(produto_key)
            
            if unidade not in grupos:
                grupos[unidade] = {
                    'dados': [],
                    'divisor': divisor
                }
            grupos[unidade]['dados'].append(row)

        # Definir altura da figura baseado no tipo de gráfico
        n_grupos = len(grupos)
        
        # Ajustar altura para gráficos únicos (maior) e múltiplos (padrão)
        if tabela_origem in tabelas_com_grafico_unico:
            figsize = (14, 10)  # Altura maior para gráficos únicos (1000px)
        else:
            figsize = (14, 6 * n_grupos)  # Altura padrão para múltiplos gráficos

        # Criar figura com o tamanho apropriado
        fig, axs = plt.subplots(n_grupos, 1, figsize=figsize)
        if n_grupos == 1:
            axs = [axs]  # Garantir que axs seja sempre uma lista

        # Extrair anos uma única vez (assumindo que todos os produtos têm os mesmos anos)
        anos_colunas = [col for col in df_municipio.columns if isinstance(col, str) and col.startswith('20')]
        anos_colunas = sorted(anos_colunas, key=lambda x: int(x[-2:]))
        anos_int = [int(ano[-2:]) for ano in anos_colunas]

        # Plotar cada grupo em um subplot
        for i, (unidade, grupo) in enumerate(grupos.items()):
            ax = axs[i]
            dados_grupo = grupo['dados']
            divisor = grupo['divisor']
            
            for row in dados_grupo:
                produto_info = row['Produto']
                produto_nome, _, cor = get_produto_info(produto_info)
                
                valores = [row[ano]/divisor if pd.notna(row[ano]) else None for ano in anos_colunas]
                
                if all(pd.isna(valores)):
                    continue
                
                # Converter para arrays numpy
                valores_arr = np.array(valores)
                anos_arr = np.array(anos_int)
                mask = ~pd.isna(valores_arr)
                
                # Plotar linha
                ax.plot(anos_arr[mask], valores_arr[mask], 
                       marker='o', 
                       linestyle='-',
                       color=cor,
                       label=produto_nome,
                       linewidth=2.5,
                       markersize=8,
                       markeredgecolor='white',
                       markeredgewidth=1)
            
            # Configurações do gráfico
            titulo_grupo = f"{titulo_base} - {municipio}" if i == 0 else ""
            ax.set_title(titulo_grupo, fontsize=16, pad=20, fontweight='bold')
            
            # Label do eixo Y personalizado
            ylabel = config.get('ylabel', tipo_dado) if i == 0 and 'ylabel' in config else tipo_dado
            ax.set_ylabel(f"{ylabel} ({unidade})", fontsize=12)
            
            ax.set_xlabel('Ano', fontsize=12)
            ax.set_xticks(anos_int)
            ax.set_xticklabels([f"20{ano}" for ano in anos_int], rotation=45 if len(anos_int) > 5 else 0)
            
            # Grid e fundo branco
            ax.grid(True, linestyle=':', alpha=0.6)
            ax.set_facecolor('white')
            
            # Legenda na parte inferior
            handles, labels = ax.get_legend_handles_labels()
            ax.legend(handles, labels,
                     loc='upper center',
                     bbox_to_anchor=(0.5, -0.15),  # Ajustado para melhor posicionamento
                     fontsize=10,
                     framealpha=0.9,
                     ncol=2)
        
        # Ajustar layout para acomodar a legenda
        plt.tight_layout()
        if tabela_origem in tabelas_com_grafico_unico:
            plt.subplots_adjust(bottom=0.25)  # Mais espaço para legenda em gráficos únicos
        else:
            plt.subplots_adjust(bottom=0.1 + 0.05 * n_grupos)  # Espaço proporcional para múltiplos gráficos
        
        fig.patch.set_facecolor('white')
        
        # Salvar para buffer
        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=150, bbox_inches='tight', facecolor=fig.get_facecolor())
        buf.seek(0)
        img = Image.open(buf)
        plt.close()
        return img
            
    except Exception as e:
        print(f"Erro ao criar gráfico para {municipio}: {str(e)}")
        plt.close()
        return None

def gerar_excel_agro(dados_agro, nome_bacia_export):
    try:
        output = BytesIO()
        workbook = Workbook()
        workbook.remove(workbook.active)
        
        graficos_por_municipio = {}
        
        def get_nome_produto(valor):
            return valor[0] if isinstance(valor, tuple) else valor
            
        for nome_tabela, dados in dados_agro.items():
            if nome_tabela == 'IBGE_Municipios_ZAP':
                ws = workbook.create_sheet(title='IBGE_Municipios')
                for r in dataframe_to_rows(dados, index=True, header=True):
                    r = [get_nome_produto(cell) if isinstance(cell, str) and cell in DICIONARIO_PRODUTOS else cell for cell in r]
                    ws.append(r)
                continue
            
            if isinstance(dados, dict):
                sheet_name = nome_tabela[:31]
                ws = workbook.create_sheet(title=sheet_name)
                current_row = 1
                
                # Determinar o tipo de dado para o título do gráfico
                tipo_dado = "Quantidade Produzida" if "Quantidade" in nome_tabela else \
                           "Valor da Produção" if "Valor" in nome_tabela else \
                           "Efetivo" if "Efetivo" in nome_tabela else "Dados"
                
                for municipio, df in dados.items():
                    if not df.empty:
                        df_display = df.copy()
                        df_display['Produto'] = df_display['Produto'].apply(get_nome_produto)
                        
                        ws.append([municipio] + ['']*(len(df_display.columns)-1))
                        ws.merge_cells(start_row=current_row, start_column=1, 
                                      end_row=current_row, end_column=len(df_display.columns))
                        cell = ws.cell(row=current_row, column=1)
                        from copy import copy
                        font = copy(cell.font)
                        font.bold = True
                        cell.font = font
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        current_row += 1
                        
                        header = ['Produto'] + [str(col)[-4:] if str(col).startswith('20') else col for col in df_display.columns[1:]]
                        ws.append(header)
                        
                        for col in range(1, len(header)+1):
                            cell = ws.cell(row=current_row, column=col)
                            font = copy(cell.font)
                            font.bold = True
                            cell.font = font
                        
                        current_row += 1
                        
                        for _, row in df_display.iterrows():
                            ws.append(row.tolist())
                            current_row += 1
                        
                        ws.append(['']*len(df_display.columns))
                        current_row += 1
                        
                        # CORREÇÃO AQUI: Passando o nome_tabela como tabela_origem
                        img = criar_grafico_unico_municipio(df, municipio, tipo_dado, nome_tabela)
                        if img:
                            graficos_por_municipio[(nome_tabela, municipio)] = img
        
        workbook.save(output)
        output.seek(0)
        
        # Exportar para o Google Drive
        try:
            drive_service = build('drive', 'v3', credentials=st.session_state["ee_credentials"])
            
            # 1. Verificar/Criar pasta ZAP
            query = "name='ZAP' and mimeType='application/vnd.google-apps.folder' and trashed=false"
            results = drive_service.files().list(q=query, fields="files(id, name)").execute()
            items = results.get('files', [])
            
            zap_folder_id = items[0]['id'] if items else None
            
            # 2. Criar subpasta para os gráficos
            subfolder_name = f"{nome_bacia_export}_graficos"
            query = f"name='{subfolder_name}' and mimeType='application/vnd.google-apps.folder' and '{zap_folder_id}' in parents and trashed=false"
            results = drive_service.files().list(q=query, fields="files(id, name)").execute()
            items = results.get('files', [])
            
            if items:
                graficos_folder_id = items[0]['id']
            else:
                file_metadata = {
                    'name': subfolder_name,
                    'mimeType': 'application/vnd.google-apps.folder',
                    'parents': [zap_folder_id]
                }
                folder = drive_service.files().create(body=file_metadata, fields='id').execute()
                graficos_folder_id = folder.get('id')
            
            # 3. Upload dos gráficos
            uploaded_graphs = 0
            for (tabela, municipio), img in graficos_por_municipio.items():
                try:
                    img_bytes = io.BytesIO()
                    img.save(img_bytes, format='PNG')
                    img_bytes.seek(0)
                    
                    nome_arquivo = f"{tabela[:20]}_{municipio[:30]}.png".replace("/", "_").replace("\\", "_")
                    
                    file_metadata = {
                        'name': nome_arquivo,
                        'parents': [graficos_folder_id],
                        'mimeType': 'image/png'
                    }
                    
                    media = MediaIoBaseUpload(img_bytes, mimetype='image/png')
                    
                    drive_service.files().create(
                        body=file_metadata,
                        media_body=media,
                        fields='id'
                    ).execute()
                    
                    uploaded_graphs += 1
                except Exception as e:
                    print(f"Erro ao enviar gráfico {nome_arquivo}: {e}")
            
            st.success(f"✅ {uploaded_graphs} gráficos salvos na pasta '{subfolder_name}' no Google Drive")
            
            # 4. Upload do arquivo Excel
            file_metadata = {
                'name': f"{nome_bacia_export}_dados_agro.xlsx",
                'parents': [zap_folder_id],
                'mimeType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            }
            
            media = MediaIoBaseUpload(output, 
                                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                                    resumable=True)
            
            drive_service.files().create(
                body=file_metadata,
                media_body=media,
                fields='id'
            ).execute()
            
        except Exception as e:
            st.error(f"❌ Erro ao exportar para o Google Drive: {str(e)}")
            print(f"Erro detalhado: {traceback.format_exc()}")
        
        return output
        
    except Exception as e:
        st.error(f"Erro ao gerar Excel: {e}")
        print(f"Erro detalhado: {traceback.format_exc()}")
        return None

# 6. Processamento principal
def process_data(geometry, crs, nome_bacia_export="bacia", process_type="all"):
    try:
        # Verificar se o Earth Engine está inicializado com o projeto correto
        if "selected_project" not in st.session_state or "ee_credentials" not in st.session_state:
            st.error("Earth Engine não foi inicializado corretamente. Por favor, reconecte-se.")
            return None
            
        # Garantir que o Earth Engine está inicializado com o projeto selecionado
        try:
            ee.Initialize(st.session_state["ee_credentials"], project=st.session_state["selected_project"])
        except Exception as e:
            st.error(f"Erro ao reinicializar Earth Engine: {e}")
            return None
        
        data_atual = datetime.datetime.now()
        mes_formatado = data_atual.strftime("%b")  # Ex: "Jan"
        ano_atual = data_atual.year
        ano_anterior = ano_atual - 1
        
        # Inicializar dicionário de resultados
        resultados = {
            "nome_bacia_export": nome_bacia_export,
            "mes_formatado": mes_formatado,
            "ano_atual": ano_atual,
            "ano_anterior": ano_anterior,
        }
        
        # Processar apenas dados agro se especificado
        if process_type == "agro":
            st.info("Processando dados agro e socioeconômicos...")
            municipios_df = processar_municipios(geometry, nome_bacia_export)
            
            if municipios_df is not None:
                geocodigos = municipios_df['geocodigo'].tolist()
                dados_agro = processar_tabelas_agro(geocodigos)
                return dados_agro
            
            return None
        
        # Processar apenas sensoriamento remoto
        elif process_type == "remoto":
            # Calcular o bounding box da geometria
            bbox = geometry.bounds()
            bacia = bbox.buffer(1000)  # 1000 metros = 1 km
            periodo_fim = ee.Date(data_atual.strftime("%Y-%m-%d"))
            periodo_inicio = periodo_fim.advance(-365, 'day')

            # Carregar imagens Sentinel-2 se selecionado
            if st.session_state.get("exportar_sentinel_composite"):
                sentinel = ee.ImageCollection("COPERNICUS/S2_SR_HARMONIZED") \
                    .select(['B1', 'B2', 'B3', 'B4', 'B5', 'B6', 'B7', 'B8', 'B8A', 'B9', 'B11', 'B12']) \
                    .filterMetadata('CLOUDY_PIXEL_PERCENTAGE', 'less_than', 10) \
                    .filterBounds(bacia) \
                    .filterDate(periodo_inicio, periodo_fim)

                num_imagens = sentinel.size().getInfo()
                if num_imagens == 0:
                    st.error("Nenhuma imagem Sentinel-2 encontrada para o período especificado.")
                else:
                    st.success(f"Imagens Sentinel-2 encontradas: {num_imagens}")
                    sentinel_median = sentinel.median().clip(bacia)
                    sentinel_composite = sentinel_median.select(['B2', 'B3', 'B4', 'B8']).rename(['B2', 'B3', 'B4', 'B8'])
                    
                    try:
                        # Criar uma FeatureCollection com as informações das imagens
                        sentinel_list = sentinel.toList(sentinel.size())
                        features = ee.FeatureCollection(sentinel_list.map(lambda img: ee.Feature(None, {
                            'id': ee.Image(img).id(),
                            'date': ee.Image(img).date().format('YYYY-MM-dd'),
                            'cloud_cover': ee.Image(img).get('CLOUDY_PIXEL_PERCENTAGE')
                        })))

                        # Exportar a lista de imagens para um arquivo CSV
                        export_task = ee.batch.Export.table.toDrive(
                            collection=features,
                            folder='zap',
                            description='lista_imagens_sentinel-2',
                            fileFormat='CSV'
                        )
                        export_task.start()
                        st.success("Exportação da lista de imagens Sentinel-2 iniciada. Verifique seu Google Drive na pasta 'zap'.")
                    except Exception as e:
                        st.error(f"Erro ao exportar a lista de imagens Sentinel-2: {e}")

            # Gerar índices se selecionados
            indices = {}
            if st.session_state.get("exportar_ndvi"):
                indices["NDVI"] = sentinel_median.normalizedDifference(['B8', 'B4'])
            if st.session_state.get("exportar_gndvi"):
                indices["GNDVI"] = sentinel_median.normalizedDifference(['B8', 'B3'])
            if st.session_state.get("exportar_ndwi"):
                indices["NDWI"] = sentinel_median.normalizedDifference(['B3', 'B8'])
            if st.session_state.get("exportar_ndmi"):
                indices["NDMI"] = sentinel_median.normalizedDifference(['B8', 'B11'])

            # Carregar MDE e Declividade se selecionados
            if st.session_state.get("exportar_srtm_mde") or st.session_state.get("exportar_declividade"):
                mde_proj = ee.ImageCollection("JAXA/ALOS/AW3D30/V3_2").filterBounds(bacia).first().select(0).projection()
                mde = ee.ImageCollection("JAXA/ALOS/AW3D30/V3_2").filterBounds(bacia).mosaic().clip(bacia).setDefaultProjection(mde_proj)
                elevation = mde.select('DSM')

                if st.session_state.get("exportar_declividade"):
                    declividade_graus = ee.Terrain.slope(elevation)
                    declividade = declividade_graus.divide(180).multiply(3.14159).tan().multiply(100)
                    declividade_reclass = declividade.expression(
                        "b(0) == 0 ? 1 : " + 
                        "b(0) <= 3 ? 1 : " + 
                        "(b(0) > 3 && b(0) <= 8) ? 2 : " + 
                        "(b(0) > 8 && b(0) <= 20) ? 3 : " + 
                        "(b(0) > 20 && b(0) <= 45) ? 4 : " + 
                        "(b(0) > 45 && b(0) <= 75) ? 5 : " + 
                        "(b(0) > 75) ? 6 : -1"
                    )

            # Carregar MapBiomas 2023 se selecionado
            if st.session_state.get("exportar_mapbiomas"):
                mapbiomas = ee.Image("projects/mapbiomas-public/assets/brazil/lulc/collection9/mapbiomas_collection90_integration_v1") \
                    .select('classification_2023') \
                    .clip(bacia)

            # Carregar Qualidade de Pastagens se selecionado
            if st.session_state.get("exportar_pasture_quality"):
                pasture_quality = ee.Image("projects/mapbiomas-public/assets/brazil/lulc/collection9/mapbiomas_collection90_pasture_quality_v1") \
                    .select('pasture_quality_2023') \
                    .clip(bacia)

            # Carregar PUC (UFV, IBGE, Embrapa) se selecionados
            if st.session_state.get("exportar_puc_ufv"):
                puc_ufv = ee.ImageCollection('users/zap/puc_ufv').filterBounds(bacia).mosaic().clip(bacia)
            if st.session_state.get("exportar_puc_ibge"):
                puc_ibge = ee.ImageCollection('users/zap/puc_ibge').filterBounds(bacia).mosaic().clip(bacia)
            if st.session_state.get("exportar_puc_embrapa"):
                puc_embrapa = ee.ImageCollection('users/zap/puc_embrapa').filterBounds(bacia).mosaic().clip(bacia)

            # Carregar Landforms se selecionado
            if st.session_state.get("exportar_landforms"):
                landforms = ee.Image('CSP/ERGo/1_0/Global/SRTM_landforms').clip(bacia)

            # Determinar o EPSG com base no fuso
            fusos_mg = ee.FeatureCollection('users/zap/fusos_mg')
            fuso_maior_area = fusos_mg.filterBounds(bacia).map(lambda f: f.set('area', f.area())).sort('area', False).first()
            epsg = fuso_maior_area.get('epsg').getInfo()

            # Reprojetar todas as imagens selecionadas
            if st.session_state.get("exportar_srtm_mde"):
                resultados["utm_elevation"] = reprojetarImagem(elevation, epsg, 30)
            if st.session_state.get("exportar_declividade"):
                resultados["utm_declividade"] = reprojetarImagem(declividade_reclass, epsg, 30).float()
            if st.session_state.get("exportar_ndvi"):
                resultados["utm_ndvi"] = reprojetarImagem(indices["NDVI"], epsg, 10)
            if st.session_state.get("exportar_gndvi"):
                resultados["utm_gndvi"] = reprojetarImagem(indices["GNDVI"], epsg, 10)
            if st.session_state.get("exportar_ndwi"):
                resultados["utm_ndwi"] = reprojetarImagem(indices["NDWI"], epsg, 10)
            if st.session_state.get("exportar_ndmi"):
                resultados["utm_ndmi"] = reprojetarImagem(indices["NDMI"], epsg, 10)
            if st.session_state.get("exportar_sentinel_composite"):
                resultados["utm_sentinel2"] = reprojetarImagem(sentinel_composite, epsg, 10).float()
            if st.session_state.get("exportar_mapbiomas"):
                resultados["utm_mapbiomas"] = reprojetarImagem(mapbiomas, epsg, 30)
            if st.session_state.get("exportar_pasture_quality"):
                resultados["utm_pasture_quality"] = reprojetarImagem(pasture_quality, epsg, 30).float()
            if st.session_state.get("exportar_landforms"):
                resultados["utm_landforms"] = reprojetarImagem(landforms, epsg, 30)
            if st.session_state.get("exportar_puc_ufv"):
                resultados["utm_puc_ufv"] = reprojetarImagem(puc_ufv, epsg, 30).float()
            if st.session_state.get("exportar_puc_ibge"):
                resultados["utm_puc_ibge"] = reprojetarImagem(puc_ibge, epsg, 30).float()
            if st.session_state.get("exportar_puc_embrapa"):
                resultados["utm_puc_embrapa"] = reprojetarImagem(puc_embrapa, epsg, 30).float()
            
            return resultados
        
        return None
        
    except Exception as e:
        st.error(f"Erro ao processar dados: {e}")
        return None

# 7. Interface do usuário
if 'token' not in st.session_state:
    st.write("Para começar, conecte-se à sua conta Google:")
    result = oauth2.authorize_button(
        "🔵 Conectar com Google",
        REDIRECT_URI, 
        SCOPE,
        icon="https://www.google.com/favicon.ico"
    )
    if result and 'token' in result:
        st.session_state.token = result.get('token')
        st.rerun()
else:
    token = st.session_state['token']
    st.success("Você está conectado à sua conta Google!")

    # Verificar se já temos credenciais e projeto inicializados
    if "ee_credentials" not in st.session_state:
        try:
            credentials = Credentials(
                token=token['access_token'],
                refresh_token=token.get('refresh_token'),
                token_uri=TOKEN_URL,
                client_id=CLIENT_ID,
                client_secret=CLIENT_SECRET,
                scopes=SCOPES
            )
            st.session_state["ee_credentials"] = credentials
            
            # Obter lista de projetos
            service = build('cloudresourcemanager', 'v1', credentials=credentials)
            projects = service.projects().list().execute().get('projects', [])
            project_ids = [project['projectId'] for project in projects]
            
            if not project_ids:
                st.warning("Nenhum projeto encontrado na sua conta do Google Cloud.")
                st.stop()
                
            st.session_state["available_projects"] = project_ids
            
            # Verificar se já temos um projeto válido armazenado
            if "selected_project" in st.session_state:
                try:
                    # Testar se o projeto armazenado ainda é válido
                    ee.Initialize(credentials, project=st.session_state["selected_project"])
                    st.session_state["ee_initialized"] = True
                    st.success(f"Earth Engine reinicializado no projeto: {st.session_state['selected_project']}")
                except Exception as e:
                    st.warning(f"Projeto anterior {st.session_state['selected_project']} não está mais disponível. Procurando novo projeto...")
                    del st.session_state["selected_project"]
            
            # Se não temos um projeto válido, procurar um que funcione
            if "selected_project" not in st.session_state:
                # Ordenar projetos por nome para consistência
                project_ids_sorted = sorted(project_ids)
                
                # Procurar um projeto com EE ativado
                selected_project = None
                for project in project_ids_sorted:
                    try:
                        ee.Initialize(credentials, project=project)
                        selected_project = project
                        break
                    except Exception as e:
                        continue
                
                if selected_project:
                    st.session_state["selected_project"] = selected_project
                    st.session_state["ee_initialized"] = True
                    st.success(f"Earth Engine inicializado com sucesso no projeto: {selected_project}")
                else:
                    # Se nenhum projeto tiver EE ativado, mostrar seleção manual
                    st.warning("Nenhum projeto com Earth Engine ativado foi encontrado automaticamente.")
                    selected_project = st.selectbox(
                        "Selecione manualmente um projeto com Earth Engine ativado:", 
                        project_ids_sorted,
                        key="project_selection"
                    )
                    if st.button("Confirmar Projeto"):
                        try:
                            ee.Initialize(credentials, project=selected_project)
                            st.session_state["selected_project"] = selected_project
                            st.session_state["ee_initialized"] = True
                            st.success(f"Earth Engine inicializado com sucesso no projeto: {selected_project}")
                            st.rerun()
                        except Exception as e:
                            st.error(f"Erro ao inicializar Earth Engine: {e}. Verifique se a API está ativada para este projeto.")
                            st.stop()
            
        except Exception as e:
            st.error(f"Erro ao inicializar o Earth Engine: {e}")
            st.stop()

    # Verificar se há processamento em andamento
    if st.session_state.get("processing"):
        st.warning("Processamento em andamento...")
        
        if st.session_state.get("uploaded_file_name"):
            st.write(f"Arquivo: {st.session_state.uploaded_file_name}")
        
        if st.button("❌ Cancelar Processamento"):
            st.session_state.processing = False
            st.session_state.processing_stage = None
            st.rerun()
        
        # Mostrar status atual
        if st.session_state.get("processing_stage") == 'agro':
            st.info("Processando dados agropecuários...")
        elif st.session_state.get("processing_stage") == 'remoto':
            completed = sum(1 for task in st.session_state.get("tasks_remoto", []) if task.status()["state"] == "COMPLETED")
            total = len(st.session_state.get("tasks_remoto", []))
            
            st.progress(completed / total)
            st.write(f"Exportando imagens: {completed}/{total} concluído")
            st.info(f"Acompanhe o progresso em: https://code.earthengine.google.com/tasks")
        
        st.stop()

    if st.session_state.get("ee_initialized"):
        uploaded_file = st.file_uploader(
            "Carregue o arquivo GeoJSON da bacia (apenas 1 polígono/multipolígono, SIRGAS 2000 (4674), máximo 1 MB)",
            type=["geojson"],
            accept_multiple_files=False,
            help="Seu arquivo tem de estar projetado em SIRGAS 2000 (4674). Use ferramentas como QGIS ou geojson.io para garantir que seu arquivo tem apenas UMA geometria"
        )
        if uploaded_file is not None:
            geometry, crs = load_geojson(uploaded_file)
            if geometry:
                nome_bacia_export = st.text_input(
                    "Digite o nome para exportação (sem espaços ou caracteres especiais). "
                    "Esse nome deve seguir o padrão utilizado para todos os produtos SIG do ZAP "
                    "(Ex.: Para o Ribeirão Santa Juliana foi utilizado o nome Rib_Santa_Juliana):",
                    placeholder="Ex: Rib_Santa_Juliana",
                    help="⚠️ Este campo é obrigatório e deve seguir o padrão de nomenclatura do ZAP"
                )
                
                if nome_bacia_export:
                    st.session_state.uploaded_file_name = uploaded_file.name
                    
                    col1, col2 = st.columns([4,1])
                    with col2:
                        if st.button("✅ Marcar Todos"):
                            st.session_state.select_all = not st.session_state.get('select_all', False)
                            st.session_state.select_ibge = st.session_state.select_all
                            st.rerun()
                    
                    with st.form(key='product_selection_form'):
                        st.subheader("📡 Produtos de Sensoriamento Remoto (Imagens/Raster)")
                        st.caption(f"Sistema de referência espacial: {crs}")
                        
                        col1, col2 = st.columns(2)
                        with col1:
                            st.markdown("**Índices Espectrais**")
                            exportar_ndvi = st.checkbox("NDVI (10m)", value=st.session_state.get('select_all', False))
                            exportar_gndvi = st.checkbox("GNDVI (10m)", value=st.session_state.get('select_all', False))
                            exportar_ndwi = st.checkbox("NDWI (10m)", value=st.session_state.get('select_all', False))
                            exportar_ndmi = st.checkbox("NDMI (10m)", value=st.session_state.get('select_all', False))
                            
                            st.markdown("**Modelo Digital de Elevação**")
                            exportar_srtm_mde = st.checkbox("SRTM MDE (30m)", value=st.session_state.get('select_all', False))
                            exportar_declividade = st.checkbox("Declividade (30m)", value=st.session_state.get('select_all', False))

                        with col2:
                            st.markdown("**Cobertura e Uso da Terra**")
                            exportar_mapbiomas = st.checkbox("MapBiomas 2023 (30m)", value=st.session_state.get('select_all', False))
                            exportar_pasture_quality = st.checkbox("Qualidade de Pastagem 2023 (30m)", value=st.session_state.get('select_all', False))
                            exportar_sentinel_composite = st.checkbox("Sentinel-2 B2/B3/B4/B8 (10m)", value=st.session_state.get('select_all', False))
                            
                            st.markdown("**Potencial de Uso**")
                            exportar_puc_ufv = st.checkbox("PUC UFV (30m)", value=st.session_state.get('select_all', False))
                            exportar_puc_ibge = st.checkbox("PUC IBGE (30m)", value=st.session_state.get('select_all', False))
                            exportar_puc_embrapa = st.checkbox("PUC Embrapa (30m)", value=st.session_state.get('select_all', False))
                            
                            st.markdown("**Geomorfologia**")
                            exportar_landforms = st.checkbox("Landforms (30m)", value=st.session_state.get('select_all', False))
                        
                        st.markdown("---")
                        
                        st.subheader("📊 Dados Agro e Socioeconômicos")
                        st.caption("Municípios com representatividade >20% na bacia hidrográfica")
                        exportar_dados_agro = st.checkbox("Ativar processamento de dados do IBGE", value=st.session_state.get('select_ibge', False))
                        
                        st.markdown("---")                    
                        submit_button = st.form_submit_button(label='✅ Confirmar Seleção')

                    if submit_button:
                        st.session_state.update({
                            "exportar_srtm_mde": exportar_srtm_mde,
                            "exportar_declividade": exportar_declividade,
                            "exportar_ndvi": exportar_ndvi,
                            "exportar_gndvi": exportar_gndvi,
                            "exportar_ndwi": exportar_ndwi,
                            "exportar_ndmi": exportar_ndmi,
                            "exportar_mapbiomas": exportar_mapbiomas,
                            "exportar_pasture_quality": exportar_pasture_quality,
                            "exportar_sentinel_composite": exportar_sentinel_composite,
                            "exportar_puc_ufv": exportar_puc_ufv,
                            "exportar_puc_ibge": exportar_puc_ibge,
                            "exportar_puc_embrapa": exportar_puc_embrapa,
                            "exportar_landforms": exportar_landforms,
                            "exportar_dados_agro": exportar_dados_agro
                        })
                        st.success("Seleção de produtos confirmada!")
                                    
                    if nome_bacia_export:  # Verifica primeiro se o nome foi preenchido
                        if st.session_state.get("exportar_srtm_mde") is not None:
                            if not st.session_state.processing:
                                if st.button("Processar Dados"):
                                    st.session_state.processing = True
                                    st.session_state.tasks_remoto = []
                                    st.session_state.completed_tasks = 0
                                    st.session_state.resultados = None
                                    st.session_state.agro_completed = False
                                    st.rerun()
                                            
                        if st.session_state.processing:
                            with st.spinner("Processando dados, por favor aguarde..."):
                                # Verificar se deve processar sensoriamento remoto
                                process_remoto = any([
                                    st.session_state.get("exportar_srtm_mde"),
                                    st.session_state.get("exportar_declividade"), 
                                    st.session_state.get("exportar_ndvi"),
                                    st.session_state.get("exportar_gndvi"),
                                    st.session_state.get("exportar_ndwi"),
                                    st.session_state.get("exportar_ndmi"),
                                    st.session_state.get("exportar_sentinel_composite"),
                                    st.session_state.get("exportar_mapbiomas"),
                                    st.session_state.get("exportar_pasture_quality"),
                                    st.session_state.get("exportar_landforms"),
                                    st.session_state.get("exportar_puc_ufv"),
                                    st.session_state.get("exportar_puc_ibge"),
                                    st.session_state.get("exportar_puc_embrapa")
                                ])
                                
                                # Verificar se deve processar dados agro
                                st.session_state.process_agro = st.session_state.get("exportar_dados_agro")
                                
                                # Processar primeiro os dados agro, se selecionados
                                if st.session_state.process_agro and not st.session_state.agro_completed:
                                    st.session_state.processing_stage = 'agro'
                                    
                                    if not st.session_state.tasks_agro:  # Primeira execução
                                        st.info("Processando dados agropecuários e socioeconômicos...")
                                        municipios_df = processar_municipios(geometry, nome_bacia_export)
                                        if municipios_df is not None:
                                            geocodigos = [int(x) for x in municipios_df['geocodigo'].tolist()]
                                            dados_agro = processar_tabelas_agro(geocodigos)
                                            if dados_agro:
                                                excel_agro = gerar_excel_agro(dados_agro, nome_bacia_export)
                                                if excel_agro:
                                                    st.session_state.tasks_agro = [excel_agro]
                                                    st.session_state.agro_completed = True
                                                    
                                                    st.download_button(
                                                        label="📥 Baixar Dados Agro",
                                                        data=excel_agro,
                                                        file_name=f"{nome_bacia_export}_dados_agro.xlsx",
                                                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                                                    )
                                                    st.success("✅ Dados agro exportados para o Google Drive!")
                                    
                                    st.rerun()
                                
                                # Processar sensoriamento remoto depois
                                elif process_remoto and (st.session_state.agro_completed or not st.session_state.process_agro):
                                    st.session_state.processing_stage = 'remoto'
                                    
                                    if st.session_state.resultados is None:
                                        st.info("Preparando imagens de sensoriamento remoto...")
                                        st.session_state.resultados = process_data(geometry, crs, nome_bacia_export, "remoto")
                                    
                                    if st.session_state.resultados and not st.session_state.tasks_remoto:
                                        resultados = st.session_state.resultados
                                        tasks = []
                                        
                                        if st.session_state.get("exportar_srtm_mde") and "utm_elevation" in resultados:
                                            tasks.append(exportarImagem(resultados["utm_elevation"], "06_", "_SRTM_MDE", 30, geometry, nome_bacia_export))
                                        if st.session_state.get("exportar_declividade") and "utm_declividade" in resultados:
                                            tasks.append(exportarImagem(resultados["utm_declividade"], "02_", "_Declividade", 30, geometry, nome_bacia_export))
                                        if st.session_state.get("exportar_ndvi") and "utm_ndvi" in resultados:
                                            tasks.append(exportarImagem(resultados["utm_ndvi"], "06_", f"_NDVImediana_{resultados['mes_formatado']}{resultados['ano_anterior']}-{resultados['ano_atual']}", 10, geometry, nome_bacia_export))
                                        if st.session_state.get("exportar_gndvi") and "utm_gndvi" in resultados:
                                            tasks.append(exportarImagem(resultados["utm_gndvi"], "06_", f"_GNDVI_{resultados['mes_formatado']}{resultados['ano_anterior']}-{resultados['ano_atual']}", 10, geometry, nome_bacia_export))
                                        if st.session_state.get("exportar_ndwi") and "utm_ndwi" in resultados:
                                            tasks.append(exportarImagem(resultados["utm_ndwi"], "06_", f"_NDWI_{resultados['mes_formatado']}{resultados['ano_anterior']}-{resultados['ano_atual']}", 10, geometry, nome_bacia_export))
                                        if st.session_state.get("exportar_ndmi") and "utm_ndmi" in resultados:
                                            tasks.append(exportarImagem(resultados["utm_ndmi"], "06_", f"_NDMI_{resultados['mes_formatado']}{resultados['ano_anterior']}-{resultados['ano_atual']}", 10, geometry, nome_bacia_export))
                                        if st.session_state.get("exportar_sentinel_composite") and "utm_sentinel2" in resultados:
                                            tasks.append(exportarImagem(resultados["utm_sentinel2"], "06_", f"_S2_B2B3B4B8_{resultados['mes_formatado']}{resultados['ano_anterior']}-{resultados['ano_atual']}", 10, geometry, nome_bacia_export))
                                        if st.session_state.get("exportar_mapbiomas") and "utm_mapbiomas" in resultados:
                                            tasks.append(exportarImagem(resultados["utm_mapbiomas"], "06_", "_MapBiomas_col9_2023", 30, geometry, nome_bacia_export))
                                        if st.session_state.get("exportar_pasture_quality") and "utm_pasture_quality" in resultados:
                                            tasks.append(exportarImagem(resultados["utm_pasture_quality"], "06_", "_Vigor_Pastagem_col9_2023", 30, geometry, nome_bacia_export))
                                        if st.session_state.get("exportar_landforms") and "utm_landforms" in resultados:
                                            tasks.append(exportarImagem(resultados["utm_landforms"], "06_", "_Landforms", 30, geometry, nome_bacia_export))
                                        if st.session_state.get("exportar_puc_ufv") and "utm_puc_ufv" in resultados:
                                            tasks.append(exportarImagem(resultados["utm_puc_ufv"], "02_", "_PUC_UFV", 30, geometry, nome_bacia_export))
                                        if st.session_state.get("exportar_puc_ibge") and "utm_puc_ibge" in resultados:
                                            tasks.append(exportarImagem(resultados["utm_puc_ibge"], "02_", "_PUC_IBGE", 30, geometry, nome_bacia_export))
                                        if st.session_state.get("exportar_puc_embrapa") and "utm_puc_embrapa" in resultados:
                                            tasks.append(exportarImagem(resultados["utm_puc_embrapa"], "02_", "_PUC_Embrapa", 30, geometry, nome_bacia_export))
                                        
                                        st.session_state.tasks_remoto = tasks
                                        st.success("✅ Exportação de imagens iniciada na Earth Engine!")
                                        st.info(f"Acompanhe o progresso em: https://code.earthengine.google.com/tasks")
                                    
                                    # Verificação simplificada do status
                                    if st.session_state.tasks_remoto:
                                        completed = sum(1 for task in st.session_state.tasks_remoto if task.status()["state"] == "COMPLETED")
                                        total = len(st.session_state.tasks_remoto)
                                        
                                        st.progress(completed / total)
                                        st.write(f"Progresso: {completed}/{total} tarefas concluídas")
                                        
                                        if completed < total:
                                            time.sleep(15)  # Verifica a cada 15 segundos
                                            st.rerun()
                                        else:
                                            st.session_state.processing = False
                                            st.session_state.processing_stage = 'done'
                                            st.balloons()
                                            st.success("✅ Todos os processamentos foram concluídos!")
                                            st.markdown(
                                                f"[Abrir pasta 'zap' no Google Drive](https://drive.google.com/drive/)",
                                                unsafe_allow_html=True)
                    else:
                        st.warning("Por favor, preencha o nome para exportação antes de selecionar os produtos.")